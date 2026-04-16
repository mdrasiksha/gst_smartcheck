import xml.etree.ElementTree as ET
from datetime import datetime
import os
import re

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ai_extractor import validate_gstin_checksum


EXCEL_COLUMNS = [
    "Voucher Type",
    "Date",
    "GSTIN",
    "Taxable Value",
    "CGST",
    "SGST",
    "IGST",
    "Total Amount",
    "Source File Name",
    "Invoice Number",
    "Validation Status",
    "Confidence Score",
    "Validation",
]

NUMERIC_COLUMNS = ["Taxable Value", "CGST", "SGST", "IGST", "Total Amount", "Confidence Score"]
CURRENCY_COLUMNS = ["Taxable Value", "CGST", "SGST", "IGST", "Total Amount"]
ALL_INVOICES_COLUMNS = [
    "File Name",
    "Invoice Number",
    "Vendor Name",
    "GSTIN",
    "Date",
    "Taxable Amount",
    "CGST",
    "SGST",
    "IGST",
    "Total Amount",
    "Confidence Score",
    "Invoice Confidence",
    "Tax Confidence",
    "Total Confidence",
]
LINE_ITEMS_COLUMNS = [
    "Invoice Number",
    "Vendor Name",
    "Line Index",
    "Description",
    "Quantity",
    "Unit Price",
    "Total Price",
]

INVALID_TEXT_VALUES = {"", ".", "-", "missing", "na", "null", "n/a", "unknown", "original"}


def _is_valid_value(val: str) -> bool:
    if val is None:
        return False
    text = str(val).strip()
    if not text:
        return False
    if text.lower() in INVALID_TEXT_VALUES:
        return False
    if len(re.sub(r"[^A-Za-z0-9]", "", text)) < 2:
        return False
    if not re.search(r"[A-Za-z0-9]", text):
        return False
    return True


def _first_available(data, keys, default=None):
    for key in keys:
        if key in data and data[key] not in (None, ""):
            return data[key]
    return default


def _extract_confidence_score(data):
    overall_confidence = data.get("Overall Confidence")
    if isinstance(overall_confidence, (int, float)):
        return round(float(overall_confidence) * 100, 2)
    confidence_data = data.get("Confidence")
    if isinstance(confidence_data, dict) and confidence_data:
        return round(sum(confidence_data.values()) / len(confidence_data) * 100, 2)
    return _first_available(data, ["Confidence Score", "Confidence"], default=None)


def _normalize_date(value):
    if value in (None, ""):
        return None

    parsed = pd.to_datetime(value, errors="coerce", dayfirst=True)
    if pd.isna(parsed):
        return None
    return parsed.strftime("%Y-%m-%d")


def _prepare_row(data, status, source_file_name=None):
    is_non_gst = data.get("Invoice Type") == "Non-GST Invoice" or data.get("Validation") == "Non GST Invoice"
    taxable_value = _first_available(data, ["Taxable Amount", "Taxable Value"])
    if is_non_gst and taxable_value in (None, ""):
        taxable_value = _first_available(data, ["Final Amount", "Total"])

    cgst = _first_available(data, ["CGST Amount", "CGST"])
    sgst = _first_available(data, ["SGST Amount", "SGST"])
    igst = _first_available(data, ["IGST Amount", "IGST"])
    if is_non_gst:
        cgst, sgst, igst = 0, 0, 0

    row = {
        "Voucher Type": "Purchase",
        "Date": _normalize_date(_first_available(data, ["Invoice Date", "Date"])),
        "GSTIN": str(_first_available(data, ["GST Number", "GSTIN"], default="") or "").upper() or None,
        "Taxable Value": taxable_value,
        "CGST": cgst,
        "SGST": sgst,
        "IGST": igst,
        "Total Amount": _first_available(data, ["Final Amount", "Total"]),
        "Source File Name": os.path.basename(source_file_name or _first_available(data, ["Source File Name", "File Name", "Filename"], default="") or "") or None,
        "Invoice Number": _first_available(data, ["Invoice Number", "Invoice No"]),
        "Validation Status": status,
        "Confidence Score": _extract_confidence_score(data),
        "Validation": data.get("Validation", "Math Mismatch"),
    }

    frame = pd.DataFrame([row], columns=EXCEL_COLUMNS)

    for col in NUMERIC_COLUMNS:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")

    for col in CURRENCY_COLUMNS:
        frame[col] = frame[col].round(2)

    return frame


def _build_summary_frame(data, source_file_name=None):
    total_amount = _to_numeric(_first_available(data, ["Final Amount", "Total"]))
    return pd.DataFrame(
        [
            {
                "Generated Time": datetime.utcnow().strftime("%Y-%m-%d %H:%M:%S UTC"),
                "Total Files": 1,
                "Total Amount": round(total_amount or 0.0, 2),
            }
        ]
    )


def _to_numeric(value):
    if value in (None, ""):
        return None
    try:
        return float(str(value).replace(",", "").strip())
    except (TypeError, ValueError):
        return None


def _clean_text_value(value):
    text = "" if value is None else str(value).strip()
    if text in {"", ".", "-", "—", '""', "''"}:
        return ""
    return text


def clean_text(text):
    if text is None:
        return ""
    cleaned = str(text).replace("|", " ").replace("-", " ").replace(".", " ")
    cleaned = re.sub(r"\s+", " ", cleaned).strip()
    tokens = [token for token in cleaned.split() if len(re.sub(r"[^A-Za-z0-9]", "", token)) > 1]
    final_text = " ".join(tokens).strip()
    return final_text if _is_valid_value(final_text) else ""


def _infer_quantity_from_text(description: str):
    if not description:
        return None
    qty_match = re.search(r"\b(?:QTY|QUANTITY)\s*[:x-]?\s*(\d+(?:\.\d+)?)\b", description, flags=re.IGNORECASE)
    if qty_match:
        return _to_numeric(qty_match.group(1))
    trailing_number = re.search(r"\b(\d+(?:\.\d+)?)\b", description)
    if trailing_number:
        return _to_numeric(trailing_number.group(1))
    return None


def _extract_header_identity(data):
    invoice_number = _clean_text_value(_first_available(data, ["Invoice Number", "Invoice No"]))
    vendor_name = _clean_text_value(_first_available(data, ["Vendor Name", "Supplier Name"]))
    fallback_text = "\n".join(
        str(data.get(key) or "")
        for key in ("Raw Text", "OCR Text", "raw_text", "text")
        if data.get(key)
    )

    if not invoice_number and fallback_text:
        invoice_match = re.search(
            r"(?:INVOICE|INV|BILL)\s*(?:NO|NUMBER)?\.?\s*[:\-]?\s*([A-Z0-9][A-Z0-9\-/]*)",
            fallback_text,
            flags=re.IGNORECASE,
        )
        if invoice_match:
            invoice_number = _clean_text_value(invoice_match.group(1))

    if not vendor_name and fallback_text:
        for line in fallback_text.splitlines():
            cleaned = _clean_text_value(line)
            if not cleaned:
                continue
            if re.search(r"\b(INVOICE|BILL TO|SHIP TO|GSTIN|GST)\b", cleaned, flags=re.IGNORECASE):
                continue
            if re.search(r"[A-Za-z]{3,}", cleaned):
                vendor_name = cleaned
                break

    invoice_number = invoice_number if _is_valid_value(invoice_number) else None
    vendor_name = vendor_name if _is_valid_value(vendor_name) else None

    return {
        "Invoice Number": invoice_number,
        "Vendor Name": vendor_name,
    }


def _build_all_invoices_frame(data, source_file_name=None):
    field_conf = data.get("Field Confidence") if isinstance(data.get("Field Confidence"), dict) else {}
    tax_conf_values = [
        field_conf.get("CGST Amount"),
        field_conf.get("SGST Amount"),
        field_conf.get("IGST Amount"),
    ]
    tax_conf_values = [value for value in tax_conf_values if isinstance(value, (int, float))]
    tax_confidence = round(sum(tax_conf_values) / len(tax_conf_values) * 100, 2) if tax_conf_values else None

    header_identity = _extract_header_identity(data)
    row = {
        "File Name": os.path.basename(source_file_name or data.get("Source File Name") or ""),
        "Invoice Number": header_identity["Invoice Number"],
        "Vendor Name": header_identity["Vendor Name"],
        "GSTIN": str(_first_available(data, ["GSTIN", "GST Number"], default="") or "").upper() or None,
        "Date": _normalize_date(_first_available(data, ["Invoice Date", "Date"])),
        "Taxable Amount": _to_numeric(_first_available(data, ["Taxable Amount", "Taxable Value"])),
        "CGST": _to_numeric(_first_available(data, ["CGST Amount", "CGST"])),
        "SGST": _to_numeric(_first_available(data, ["SGST Amount", "SGST"])),
        "IGST": _to_numeric(_first_available(data, ["IGST Amount", "IGST"])),
        "Total Amount": _to_numeric(_first_available(data, ["Final Amount", "Total"])),
        "Confidence Score": _to_numeric(_extract_confidence_score(data)),
        "Invoice Confidence": round(float(field_conf.get("Invoice Number", 0)) * 100, 2) if isinstance(field_conf.get("Invoice Number"), (int, float)) else None,
        "Tax Confidence": tax_confidence,
        "Total Confidence": round(float(field_conf.get("Final Amount", 0)) * 100, 2) if isinstance(field_conf.get("Final Amount"), (int, float)) else None,
    }
    frame = pd.DataFrame([row], columns=ALL_INVOICES_COLUMNS)
    for col in ["Taxable Amount", "CGST", "SGST", "IGST", "Total Amount", "Confidence Score", "Invoice Confidence", "Tax Confidence", "Total Confidence"]:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def _line_items_from_data(data):
    raw_line_items = (
        data.get("Line Items")
        or data.get("line_items")
        or data.get("Items")
        or []
    )
    if not isinstance(raw_line_items, list):
        raw_line_items = []

    header_identity = _extract_header_identity(data)
    invoice_number = header_identity["Invoice Number"]
    vendor_name = header_identity["Vendor Name"]
    rows = []
    line_index = 1
    for item in raw_line_items:
        if not isinstance(item, dict):
            continue

        description = clean_text(_first_available(item, ["Description", "Item", "Particulars"], default=""))
        qty = _to_numeric(_first_available(item, ["Quantity", "Qty", "quantity"]))
        unit_price = _to_numeric(_first_available(item, ["Unit Price", "Rate", "unit_price"]))
        total_price = _to_numeric(_first_available(item, ["Total Price", "Amount", "Line Total", "total_price"]))

        if qty is None:
            qty = _infer_quantity_from_text(description)
        if qty is None:
            continue
        if unit_price is None:
            continue
        if total_price is None and unit_price is not None:
            total_price = qty * unit_price

        if not description:
            continue

        item_row = {
            "Invoice Number": invoice_number,
            "Vendor Name": vendor_name,
            "Line Index": line_index,
            "Description": description,
            "Quantity": qty,
            "Unit Price": round(unit_price, 2) if unit_price is not None else None,
            "Total Price": round(total_price, 2) if total_price is not None else None,
        }
        print("LINE ITEM:", item_row)
        rows.append(item_row)
        line_index += 1

    return rows


def _build_line_items_frame(data):
    rows = _line_items_from_data(data)
    frame = pd.DataFrame(rows, columns=LINE_ITEMS_COLUMNS)
    for col in ["Quantity", "Unit Price", "Total Price"]:
        frame[col] = pd.to_numeric(frame[col], errors="coerce")
    return frame


def write_to_excel(data, status, output_path, source_file_name=None):
    df = _prepare_row(data, status, source_file_name=source_file_name)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Legacy Export", index=False)

    wb = load_workbook(output_path)
    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    verified_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    mismatch_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin")
    )

    for ws in wb.worksheets:
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center_align
            cell.border = thin_border

        for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
            for cell in row:
                cell.border = thin_border

    legacy_ws = wb["Legacy Export"]
    validation_col = EXCEL_COLUMNS.index("Validation") + 1
    gst_col = EXCEL_COLUMNS.index("GSTIN") + 1
    currency_col_indexes = [EXCEL_COLUMNS.index(col) + 1 for col in CURRENCY_COLUMNS]

    for row in legacy_ws.iter_rows(min_row=2, max_row=legacy_ws.max_row):
        for col_idx in currency_col_indexes:
            row[col_idx - 1].number_format = "0.00"

        validation_cell = row[validation_col - 1]
        validation_value = str(validation_cell.value or "").strip().lower()
        validation_cell.fill = verified_fill if validation_value in {"verified", "non gst invoice"} else mismatch_fill
        validation_cell.font = Font(bold=True)

        gst_cell = row[gst_col - 1]
        if gst_cell.value:
            gst_cell.value = str(gst_cell.value).upper()
            if not validate_gstin_checksum(str(gst_cell.value)):
                gst_cell.fill = mismatch_fill

    min_widths_legacy = {
        "Voucher Type": 14,
        "Date": 14,
        "GSTIN": 18,
        "Taxable Value": 14,
        "CGST": 12,
        "SGST": 12,
        "IGST": 12,
        "Total Amount": 14,
        "Source File Name": 28,
        "Invoice Number": 14,
        "Validation Status": 24,
        "Confidence Score": 16,
        "Validation": 16,
    }
    sheet_min_widths = {
        "Legacy Export": min_widths_legacy,
    }
    for ws in wb.worksheets:
        min_widths = sheet_min_widths.get(ws.title, {})
        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            header = str(ws.cell(row=1, column=col_idx).value or "")
            col_letter = ws.cell(row=1, column=col_idx).column_letter
            for row_idx in range(1, ws.max_row + 1):
                cell_value = ws.cell(row=row_idx, column=col_idx).value
                if cell_value is not None:
                    max_length = max(max_length, len(str(cell_value)))
            ws.column_dimensions[col_letter].width = max(max_length + 3, min_widths.get(header, 12))

    try:
        wb.save(output_path)
    finally:
        wb.close()

    return bool(
        output_path
        and output_path.lower().endswith(".xlsx")
        and os.path.exists(output_path)
        and os.path.getsize(output_path) > 0
    )


def generate_tally_xml(data):
    row = _prepare_row(data, status=data.get("Validation Status", "Success")).iloc[0].to_dict()

    voucher_date = ""
    if row.get("Date"):
        voucher_date = pd.to_datetime(row["Date"]).strftime("%Y%m%d")

    taxable = float(row.get("Taxable Value") or 0)
    cgst = float(row.get("CGST") or 0)
    sgst = float(row.get("SGST") or 0)
    igst = float(row.get("IGST") or 0)
    total = float(row.get("Total Amount") or (taxable + cgst + sgst + igst))
    total_tax = round(cgst + sgst + igst, 2)

    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    request_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(request_desc, "REPORTNAME").text = "Vouchers"

    request_data = ET.SubElement(import_data, "REQUESTDATA")
    tally_message = ET.SubElement(request_data, "TALLYMESSAGE", {"xmlns:UDF": "TallyUDF"})

    voucher = ET.SubElement(
        tally_message,
        "VOUCHER",
        {"VCHTYPE": "Sales", "ACTION": "Create", "OBJVIEW": "Invoice Voucher View"},
    )

    ET.SubElement(voucher, "DATE").text = voucher_date
    ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Sales"
    ET.SubElement(voucher, "PARTYGSTIN").text = str(row.get("GSTIN") or "")
    ET.SubElement(voucher, "NARRATION").text = f"GST SmartCheck import for invoice {row.get('Invoice Number') or ''}".strip()

    party_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(party_entry, "LEDGERNAME").text = "Sundry Debtors"
    ET.SubElement(party_entry, "ISDEEMEDPOSITIVE").text = "Yes"
    ET.SubElement(party_entry, "AMOUNT").text = f"-{total:.2f}"

    sales_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(sales_entry, "LEDGERNAME").text = "Sales"
    ET.SubElement(sales_entry, "ISDEEMEDPOSITIVE").text = "No"
    ET.SubElement(sales_entry, "AMOUNT").text = f"{taxable:.2f}"

    if total_tax > 0:
        tax_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(tax_entry, "LEDGERNAME").text = "Output GST"
        ET.SubElement(tax_entry, "ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(tax_entry, "AMOUNT").text = f"{total_tax:.2f}"

    return ET.tostring(envelope, encoding="utf-8", xml_declaration=True).decode("utf-8")
