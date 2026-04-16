from pathlib import Path

from openpyxl import load_workbook

from excel_writer import write_to_excel


def test_write_to_excel_only_emits_legacy_sheet(tmp_path: Path):
    output_path = tmp_path / "invoice.xlsx"
    data = {
        "Invoice Date": "2026-03-31",
        "GST Number": "27ABCDE1234F1Z5",
        "Taxable Amount": 100.0,
        "CGST Amount": 9.0,
        "SGST Amount": 9.0,
        "IGST Amount": 0.0,
        "Final Amount": 118.0,
        "Invoice Number": "INV-001",
        "Validation": "Verified",
    }

    assert write_to_excel(data, "Success", str(output_path), source_file_name="invoice.pdf")

    workbook = load_workbook(output_path)
    try:
        assert workbook.sheetnames == ["Legacy Export"]
    finally:
        workbook.close()
