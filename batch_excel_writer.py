from datetime import datetime
import xml.etree.ElementTree as ET

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from ai_extractor import validate_gstin_checksum

BATCH_COLUMNS = [
    "Source File Name",
    "Invoice No",
    "Date",
    "GSTIN",
    "Taxable Value",
    "CGST",
    "SGST",
    "IGST",
    "Total",
    "Validation Status",
    "Confidence Score",
    "Rules Applied",
    "Output File",
]

NUMERIC_COLUMNS = ["Taxable Value", "CGST", "SGST", "IGST", "Total", "Confidence Score"]


def write_batch_summary(results, output_path):
    rows = []
    for result in results:
        row = {col: result.get(col) for col in BATCH_COLUMNS}
        rows.append(row)

    df = pd.DataFrame(rows, columns=BATCH_COLUMNS)

    for col in NUMERIC_COLUMNS:
        df[col] = pd.to_numeric(df[col], errors="coerce")

    if "Date" in df:
        df["Date"] = pd.to_datetime(df["Date"], errors="coerce", dayfirst=True)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, startrow=3)

    wb = load_workbook(output_path)
    ws = wb.active

    ws.merge_cells(f"A1:{ws.cell(row=1, column=ws.max_column).column_letter}1")
    ws["A1"] = "GST SmartCheck Audit Report (Batch Summary)"
    ws["A2"] = f"Extraction Date: {datetime.now().strftime('%d-%b-%Y %H:%M:%S')}"
    ws["A1"].font = Font(size=14, bold=True)
    ws["A2"].font = Font(italic=True)

    header_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
    success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
    error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")

    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    header_row = 4
    data_row = 5

    for cell in ws[header_row]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center_align
        cell.border = thin_border

    for row in ws.iter_rows(min_row=data_row, max_row=ws.max_row):
        for cell in row:
            cell.alignment = left_align
            cell.border = thin_border

    date_col = BATCH_COLUMNS.index("Date") + 1
    gstin_col = BATCH_COLUMNS.index("GSTIN") + 1
    status_col = BATCH_COLUMNS.index("Validation Status") + 1

    for row_idx in range(data_row, ws.max_row + 1):
        date_cell = ws.cell(row=row_idx, column=date_col)
        if date_cell.value:
            date_cell.number_format = "DD-MMM-YYYY"

        gst_cell = ws.cell(row=row_idx, column=gstin_col)
        status_cell = ws.cell(row=row_idx, column=status_col)

        if gst_cell.value and not validate_gstin_checksum(str(gst_cell.value)):
            gst_cell.fill = error_fill

        if str(status_cell.value).strip().lower() == "success":
            status_cell.fill = success_fill
            status_cell.font = Font(bold=True)

    min_widths = {
        "Source File Name": 28,
        "Invoice No": 12,
        "Date": 14,
        "GSTIN": 18,
        "Taxable Value": 14,
        "CGST": 12,
        "SGST": 12,
        "IGST": 12,
        "Total": 14,
        "Validation Status": 24,
        "Confidence Score": 16,
        "Rules Applied": 26,
        "Output File": 24,
    }

    for col_idx in range(1, ws.max_column + 1):
        header = str(ws.cell(row=header_row, column=col_idx).value or "")
        col_letter = ws.cell(row=header_row, column=col_idx).column_letter
        max_length = 0
        for row_idx in range(1, ws.max_row + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is not None:
                max_length = max(max_length, len(str(value)))
        ws.column_dimensions[col_letter].width = max(max_length + 3, min_widths.get(header, 12))

    wb.save(output_path)


def generate_tally_sales_xml(invoice_data: dict, output_path: str) -> str:
    """
    Generate Tally ERP 9 / Tally Prime compatible Sales voucher XML.

    Expected invoice_data keys include:
      - GSTIN
      - Date (DD-MM-YYYY or similar)
      - Total
      - Tax
    """
    envelope = ET.Element("ENVELOPE")
    header = ET.SubElement(envelope, "HEADER")
    ET.SubElement(header, "TALLYREQUEST").text = "Import Data"

    body = ET.SubElement(envelope, "BODY")
    import_data = ET.SubElement(body, "IMPORTDATA")
    request_desc = ET.SubElement(import_data, "REQUESTDESC")
    ET.SubElement(request_desc, "REPORTNAME").text = "Vouchers"

    request_data = ET.SubElement(import_data, "REQUESTDATA")
    tally_message = ET.SubElement(request_data, "TALLYMESSAGE", {"xmlns:UDF": "TallyUDF"})

    voucher = ET.SubElement(
        tally_message,
        "VOUCHER",
        {
            "VCHTYPE": "Sales",
            "ACTION": "Create",
            "OBJVIEW": "Invoice Voucher View",
        },
    )

    ET.SubElement(voucher, "DATE").text = str(invoice_data.get("Date", "")).replace("-", "")
    ET.SubElement(voucher, "VOUCHERTYPENAME").text = "Sales"
    ET.SubElement(voucher, "PARTYGSTIN").text = str(invoice_data.get("GSTIN", ""))
    ET.SubElement(voucher, "NARRATION").text = "Imported by GST SmartCheck"

    total = float(invoice_data.get("Total", 0) or 0)
    tax = float(invoice_data.get("Tax", 0) or 0)
    taxable = round(total - tax, 2)

    party_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(party_entry, "LEDGERNAME").text = "Sundry Debtors"
    ET.SubElement(party_entry, "ISDEEMEDPOSITIVE").text = "Yes"
    ET.SubElement(party_entry, "AMOUNT").text = f"-{total:.2f}"

    sales_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
    ET.SubElement(sales_entry, "LEDGERNAME").text = "Sales"
    ET.SubElement(sales_entry, "ISDEEMEDPOSITIVE").text = "No"
    ET.SubElement(sales_entry, "AMOUNT").text = f"{taxable:.2f}"

    if tax > 0:
        tax_entry = ET.SubElement(voucher, "ALLLEDGERENTRIES.LIST")
        ET.SubElement(tax_entry, "LEDGERNAME").text = "Output GST"
        ET.SubElement(tax_entry, "ISDEEMEDPOSITIVE").text = "No"
        ET.SubElement(tax_entry, "AMOUNT").text = f"{tax:.2f}"

    tree = ET.ElementTree(envelope)
    tree.write(output_path, encoding="utf-8", xml_declaration=True)
    return output_path
